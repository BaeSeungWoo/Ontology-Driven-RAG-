# tests/unit/pipeline/test_site_b_parser.py

import os
import pytest
import openpyxl
from unittest.mock import patch, MagicMock
from langchain_core.documents import Document


@pytest.fixture(autouse=True)
def set_upstage_key():
    with patch.dict(os.environ, {"UPSTAGE_API_KEY": "test-key"}):
        yield


@pytest.fixture
def parser():
    with patch("pipeline.adapters.site_b.DoclingParser") as MockDocling, \
         patch("pipeline.adapters.site_b.UpstageParser") as MockUpstage:
        MockDocling.return_value = MagicMock()
        MockUpstage.return_value = MagicMock()
        from pipeline.adapters.site_b import SiteBParser
        p = SiteBParser()
        p._mock_docling = MockDocling.return_value
        p._mock_upstage = MockUpstage.return_value
        yield p


def make_xlsx(tmp_path, sheet_data: dict) -> str:
    """테스트용 xlsx 파일 생성. sheet_data = {sheet_name: [[row, ...], ...]}"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    for sheet_name, rows in sheet_data.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return str(path)


class TestParseManualDispatch:
    def test_xlsx_delegates_to_parse_excel(self, parser, tmp_path):
        xlsx_path = make_xlsx(tmp_path, {"Sheet1": [["부품명", "수량"], ["볼트", 10]]})
        result = parser.parse_manual(xlsx_path)
        # Docling은 호출되지 않아야 함
        parser._mock_docling.parse.assert_not_called()

    def test_pdf_delegates_to_docling(self, parser):
        parser._mock_docling.parse.return_value = [
            Document(page_content="내용", metadata={"source": "manual.pdf"})
        ]
        parser.parse_manual("/data/manual.pdf")
        parser._mock_docling.parse.assert_called_once_with("/data/manual.pdf")

    def test_pdf_adds_site_b_metadata(self, parser):
        parser._mock_docling.parse.return_value = [
            Document(page_content="내용", metadata={"source": "manual.pdf"})
        ]
        result = parser.parse_manual("/data/manual.pdf")
        for doc in result:
            assert doc.metadata["site"] == "B"


class TestParseScanned:
    def test_delegates_to_upstage(self, parser):
        parser._mock_upstage.parse.return_value = [
            Document(page_content="OCR", metadata={"source": "scan.pdf"})
        ]
        parser.parse_scanned("/data/scan.pdf")
        parser._mock_upstage.parse.assert_called_once_with("/data/scan.pdf")

    def test_adds_site_b_metadata(self, parser):
        parser._mock_upstage.parse.return_value = [
            Document(page_content="OCR", metadata={"source": "scan.pdf"})
        ]
        result = parser.parse_scanned("/data/scan.pdf")
        for doc in result:
            assert doc.metadata["site"] == "B"


class TestParseExcel:
    def test_returns_documents(self, parser, tmp_path):
        xlsx_path = make_xlsx(tmp_path, {
            "Sheet1": [["부품명", "수량"], ["볼트", 10], ["너트", 20]]
        })
        result = parser._parse_excel(xlsx_path)
        assert len(result) == 2
        assert all(isinstance(d, Document) for d in result)

    def test_formats_row_as_key_value(self, parser, tmp_path):
        xlsx_path = make_xlsx(tmp_path, {
            "Sheet1": [["이름", "값"], ["볼트", "M10"]]
        })
        result = parser._parse_excel(xlsx_path)
        assert "이름: 볼트" in result[0].page_content
        assert "값: M10" in result[0].page_content

    def test_skips_empty_rows(self, parser, tmp_path):
        xlsx_path = make_xlsx(tmp_path, {
            "Sheet1": [["이름", "값"], [None, None], ["볼트", "M10"]]
        })
        result = parser._parse_excel(xlsx_path)
        assert len(result) == 1

    def test_includes_sheet_name_in_metadata(self, parser, tmp_path):
        xlsx_path = make_xlsx(tmp_path, {
            "사양서": [["항목", "값"], ["온도", "200℃"]]
        })
        result = parser._parse_excel(xlsx_path)
        assert result[0].metadata["sheet"] == "사양서"

    def test_includes_row_number_in_metadata(self, parser, tmp_path):
        xlsx_path = make_xlsx(tmp_path, {
            "Sheet1": [["항목"], ["첫번째"], ["두번째"]]
        })
        result = parser._parse_excel(xlsx_path)
        assert result[0].metadata["row"] == 2
        assert result[1].metadata["row"] == 3

    def test_handles_multiple_sheets(self, parser, tmp_path):
        xlsx_path = make_xlsx(tmp_path, {
            "Sheet1": [["항목"], ["값A"]],
            "Sheet2": [["항목"], ["값B"]],
        })
        result = parser._parse_excel(xlsx_path)
        assert len(result) == 2
        sheets = {doc.metadata["sheet"] for doc in result}
        assert "Sheet1" in sheets
        assert "Sheet2" in sheets

    def test_skips_sheets_with_no_headers(self, parser, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "빈시트"
        # 헤더 없음 (첫 행이 비어있음)
        ws.append([None, None])
        ws.append(["값1", "값2"])
        path = tmp_path / "no_header.xlsx"
        wb.save(str(path))
        result = parser._parse_excel(str(path))
        assert result == []

    def test_sets_doc_type_spec(self, parser, tmp_path):
        xlsx_path = make_xlsx(tmp_path, {
            "Sheet1": [["항목", "값"], ["온도", "200℃"]]
        })
        result = parser._parse_excel(xlsx_path)
        for doc in result:
            assert doc.metadata["doc_type"] == "spec"

    def test_sets_site_b(self, parser, tmp_path):
        xlsx_path = make_xlsx(tmp_path, {
            "Sheet1": [["항목", "값"], ["온도", "200℃"]]
        })
        result = parser._parse_excel(xlsx_path)
        for doc in result:
            assert doc.metadata["site"] == "B"
